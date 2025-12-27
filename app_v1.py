import time
from metadata import PROJECT_METADATA as metadata
from collections import defaultdict
import psutil
from tkinter import (
    Scrollbar,
    Listbox,
    PhotoImage,
    RIGHT,
    StringVar,
    Label,
    Button,
    LEFT,
    Y,
    BOTH,
    END,
    Tk,
    Menu,
    mainloop,
    font,
    Frame,
    PhotoImage,
    filedialog,
    messagebox,
)
from project_utils import plot_graph
import openpyxl


class MotionSensorApp:
    def __init__(self, root):
        self.root = root
        self.root.title(metadata["Name"])
        self.icon = PhotoImage(file="./assets/graph.png")
        self.root.iconphoto(True, self.icon)
        self.root.resizable(False, False)

        self.window_width = 800
        self.window_height = 600
        self.center_window(self.window_width, self.window_height)

        # Create a frame to hold widgets
        self.frame = Frame(root)
        self.frame.pack(fill=BOTH, expand=True)

        # create the menu
        self.menu = Menu(self.root, tearoff=0)  # tearoff=0 to disable dashed line
        filemenu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="File", menu=filemenu)

        filemenu.add_command(label="Save", command=self.save)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=root.quit)

        helpmenu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="Help", menu=helpmenu)
        helpmenu.add_command(label="About", command=self.about)

        self.root.config(menu=self.menu)
        self.on_off_frame = Frame(self.frame)
        self.main_frame = Frame(self.frame)

        self.label_props = {
            "relief": "flat",
            "compound": "top",
            "fg": "blue",  # Text color
            "activeforeground": "blue",
            # "font": ("Arial", 20, "bold"),
            "font": font.Font(
                family="Helvetica", size=12, weight="bold", underline=True
            ),
        }

        self.tab_list = [
            {"name": "plot", "icon": PhotoImage(file="./assets/graph.png")},
            {
                "name": "performance",
                "icon": PhotoImage(file="./assets/performance.png"),
            },
            {"name": "status", "icon": PhotoImage(file="./assets/status.png")},
        ]
        # dictionary to store properties monitored
        self.properties_monitored = dict()
        # data storage
        self.data_storage = defaultdict(list)

        # image references
        self.light_off_img = PhotoImage(file="./assets/light_off_md.png")
        self.light_on_img = PhotoImage(file="./assets/light_on_md.png")
        self.power_img = PhotoImage(file="./assets/power_btn_lg.png")
        self.power_img_sm = PhotoImage(file="./assets/power_btn.png")

        # Variables
        # allows easy access to the light status
        self.light_status = StringVar()
        self.light_status.set("LIGHT ON")

        self.date_time_var = StringVar()
        self.date_time_var.set("06:12:2025 00:00:00")

        self.target_property_var = StringVar()
        self.target_property_var.set("Performance")
        # bind the target_property_var to the handle_active_tab method, so the method is called when the variable changes
        self.target_property_var.trace_add("write", self.handle_active_tab)

        # frames
        self.tab_properties_frame = None

        # build all the screens
        self.build_on_off_screen(self.on_off_frame)
        self.build_main_screen(self.main_frame)

        # show the on_off_screen
        self.show_screen(self.on_off_frame)
        # self.update_properties()

        # start date_time_update
        self.root.after(1000, self.date_time_update)  # call again in 1 second

    def about(self):
        messagebox.showinfo(
            f"About {metadata['Name']}",
            f"{metadata['description']}\n\nVersion: {metadata['Version']}\nAuthors: {', '.join(metadata['Authors'])}\nLicense: {metadata['License']}\nGitHub: {metadata['GitHub URL']}\nIssues: {metadata['Issues']}",
        )

    def save(self):
        # Ask user where to save
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="motion_sensor_data.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Measurements"

                keys, values = zip(*self.data_storage.items())
                # write header
                for col_idx, header in enumerate(keys, start=1):
                    ws.cell(row=1, column=col_idx, value=header)

                # assume all columns have same number of rows as values[0]
                num_rows = len(values[0])

                # write data
                for row_idx in range(num_rows):
                    for col_idx, column_values in enumerate(values, start=1):
                        ws.cell(
                            row=row_idx + 2,
                            column=col_idx,
                            value=column_values[row_idx],
                        )
                wb.save(file_path)
                messagebox.showinfo("Success", f"Data saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file:\n{e}")

    def show_screen(self, new_frame):
        if new_frame == self.on_off_frame:
            self.main_frame.pack_forget()
            self.on_off_frame.pack(fill=BOTH, expand=True)
        elif new_frame == self.main_frame:
            self.on_off_frame.pack_forget()
            self.main_frame.pack(fill=BOTH, expand=True)

    def get_system_usage(self):
        """
        Get the system usage
        Args:
            None
        Returns:
            dict: The system usage with the format:
                {
                    "key": "used;total;unit",
                    "cpu": "used;100%;%",
                }
        """
        # CPU usage
        cpu_percent = psutil.cpu_percent(
            interval=0
        )  # interval=0 for instant usage (non blocking main thread)
        # cpu_freq = psutil.cpu_freq()

        # RAM usage
        ram = psutil.virtual_memory()

        # Disk usage (root partition)
        disk = psutil.disk_usage("/")

        return {
            # "ram": f"{round(ram.used / (1024**3), 2)}/{round(ram.total / (1024**3), 2)}",
            "RAM": f"{round(ram.used / (1024**3), 2)};{round(ram.total / (1024**3), 2)};GB",
            # "Memory": f"{round(disk.used / (1024**3), 2)}/{round(disk.total / (1024**3), 2)}",
            "DISK": f"{round(disk.used / (1024**3), 2)};{round(disk.total / (1024**3), 2)};GB",
            "CPU": f"{cpu_percent};100;%",
        }

    def build_on_off_screen(self, frame):
        frame = Frame(frame)
        frame.pack(fill=BOTH, expand=True)

        btn = Button(
            frame,
            image=self.power_img,
            command=lambda: self.show_screen(self.main_frame),
            bd=0,
            highlightthickness=0,
            relief="flat",
            text="POWER ON",
            bg=frame.cget("bg"),  # match parent's bg color
            activebackground=frame.cget("bg"),
            cursor="hand2",
            compound="top",
            pady=15,
            fg="blue",  # Text color
            activeforeground="blue",
            font=("Arial", 22, "bold"),
        )
        btn.pack(expand=True)

    def build_main_screen(self, frame):
        # Add widgets in a grid
        left_frame = Frame(frame)

        right_frame = Frame(frame, bg="white")

        # left side widgets
        left_frame.place(relx=0, rely=0, relwidth=0.4, relheight=1)
        # left_frame.place(relx=0, rely=0, relwidth=0.395, relheight=1)
        right_frame.place(relx=0.4, rely=0, relwidth=0.6, relheight=1)

        left_top_frame = Frame(left_frame)
        left_top_frame.place(relx=0, rely=0, relwidth=1, relheight=0.25)

        # timer
        time_label = Label(
            left_top_frame,
            textvariable=self.date_time_var,
            font=("Arial", 12, "bold"),
            fg="blue",  # Text color
        )
        time_label.pack(expand=True, fill="x")

        tabs_frame = Frame(left_top_frame)
        tabs_frame.pack(expand=True)

        for index, tab in enumerate(self.tab_list):
            _name = tab["name"].capitalize()
            btn = Button(
                tabs_frame,
                text=_name,
                image=tab["icon"],
                relief="raised",
                activebackground=self.frame.cget("bg"),
                cursor="hand2",
                compound="top",
                command=lambda x=_name: self.target_property_var.set(x),
                fg="blue",  # Text color
                activeforeground="blue",
                # bd=0,
                # font=("Arial", 16, "normal"),
                font=font.Font(family="Helvetica", size=12, weight="bold"),
            )
            btn.grid(row=1, column=index, padx=5, pady=5, sticky="nsew")

            tabs_frame.grid_columnconfigure(index, weight=1, minsize=100)
        tabs_frame.grid_rowconfigure(1, weight=1, minsize=70)

        light_frame = Frame(left_frame)
        light_frame.pack(expand=True)

        bulb_label = Label(
            light_frame,
            image=self.light_on_img,
            height=300,
            textvariable=self.light_status,
            relief="flat",
            activebackground=self.frame.cget("bg"),
            compound="top",
            fg="blue",  # Text color
            activeforeground="blue",
            font=("Arial", 20, "bold"),
        )
        bulb_label.pack(expand=True)

        # right side widgets
        power_frame = Frame(right_frame, bg=right_frame.cget("bg"))
        power_frame.place(relx=0, rely=0, relwidth=1, relheight=0.2)

        power_off_btn = Button(
            power_frame,
            image=self.power_img_sm,
            command=lambda: self.show_screen(self.on_off_frame),
            relief="flat",
            highlightthickness=0,
            activebackground=right_frame.cget("bg"),
            bg=right_frame.cget("bg"),
            cursor="hand2",
            compound="top",
            bd=0,
            fg="blue",  # Text color
            activeforeground="blue",
        )
        power_off_btn.pack(
            anchor="ne",
            padx=10,
            pady=10,
        )

        target_property_label = Label(
            power_frame,
            textvariable=self.target_property_var,
            bg=right_frame.cget("bg"),
            **self.label_props,
        )
        target_property_label.pack(expand=True, anchor="center")

        self.tab_properties_frame = Frame(right_frame, bg=right_frame.cget("bg"))
        self.tab_properties_frame.place(relx=0, rely=0.2, relwidth=1, relheight=0.8)

        performance_frame = self.get_performance_frame(self.tab_properties_frame)
        performance_frame.frame_id = "performance"
        plot_frame = self.get_plot_frame(
            self.tab_properties_frame,
        )
        plot_frame.frame_id = "plot"
        status_frame = self.get_status_frame(self.tab_properties_frame)
        status_frame.frame_id = "status"

        self.handle_active_tab()

    def get_performance_frame(self, parent):
        metrics_frame = Frame(parent, bg=parent.cget("bg"))
        metrics_frame.pack(expand=True, anchor="n")
        # add the properties
        system_usage = self.get_system_usage()
        cols = 2
        for index, (key, value) in enumerate(system_usage.items()):
            row = index // cols
            col = index % cols
            property = self.create_performance_widget(metrics_frame, key, value)
            property.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)

            # metrics_frame.grid_configure(row=row, column=col, weight=1)
            metrics_frame.grid_columnconfigure(col, weight=1, minsize=150)
            metrics_frame.grid_rowconfigure(row, weight=1, minsize=100)
        return metrics_frame

    def get_plot_frame(self, parent):
        return plot_graph(
            parent,
            x_label="CPU Usage",
            y_label="Time (Seconds)",
            # values=self.data_storage["CPU"],
        )

    def get_status_frame(self, parent):
        frame = Frame(parent, bg=parent.cget("bg"))
        status_label = Label(
            frame, text="Hello from Status", bg=parent.cget("bg"), **self.label_props
        )
        status_label.pack(expand=True, anchor="n")
        return frame

    def create_performance_widget(self, widget_parent, label, value):
        """
        Create a performance widget at runtime
        Args:
            label (str): The label of the performance widget
            value (str): The value of the performance widget
        Returns:
            Frame: The performance widget
            children Widget:
                performance_label (Label): The label of the performance widget
                performance_value (Label): The value of the performance widget
        """
        # create a frame at runtime
        frame = Frame(widget_parent)
        # Set background color to white
        frame.configure(bd=5, relief="raised")
        # create label and make it a child of the frame
        performance_label = Label(frame, text=label.upper())
        performance_label.pack(expand=True, anchor="center")
        performance_label.configure(
            fg="blue",  # Text color
            font=("Arial", 10, "bold"),
        )
        # create performance value and make it a child of the layout
        performance_value = Label(frame, text=str(value))
        performance_value.pack(expand=True, fill="both", anchor="center")
        performance_value.configure(
            fg="green",  # Text color
            font=("Arial", 14, "normal"),
        )
        # add performance value to the properties_monitored dictionary
        self.properties_monitored[label] = performance_value

        return frame

    def update_properties(self):
        """
        Update the properties
        Args:
            None
        Returns:
            None
        """
        system_usage = self.get_system_usage()
        # add time stamp to the data storage
        self.data_storage["Timestamp"].append(time.strftime("%d-%m-%Y %H:%M:%S"))
        for key, value in system_usage.items():
            used, total, unit = value.split(";")
            # update the data storage
            self.data_storage[key].append(f"{used} {unit}")

            # update the performance widget
            performance = self.properties_monitored[key]
            performance["text"] = f"{used} {unit}"
            # calculate the percentage used
            percent_used = (float(used) / float(total)) * 100
            very_good_condition = percent_used <= 30
            good_condition = percent_used > 30 and percent_used <= 69
            performance.configure(
                fg="green"
                if very_good_condition
                else "blue"
                if good_condition
                else "red"
            )

    def date_time_update(self):
        self.date_time_var.set(time.strftime("%d-%m-%Y %H:%M:%S"))
        self.update_properties()
        self.root.after(1000, self.date_time_update)

    def handle_active_tab(self, *args):
        target_prop = self.target_property_var.get().lower()
        children = self.tab_properties_frame.winfo_children()
        for child in children:
            if child.frame_id == target_prop:
                child.pack(expand=True, anchor="n")
            else:
                child.pack_forget()

    def center_window(self, width, height):
        # self.root.update_idletasks()  # Important!
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Calculate position
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)

        # Apply geometry
        self.root.geometry(f"{width}x{height}+{x}+{y}")


if __name__ == "__main__":
    root = Tk()
    # create the application
    app = MotionSensorApp(root)
    # start the program
    root.mainloop()
